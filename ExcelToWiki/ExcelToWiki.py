'''
Created on Jul 21, 2015

@author: venkman69@yahoo.com
'''
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.xml.functions import fromstring,QName
from xml.etree.ElementTree import Element
import colorsys
from openpyxl.worksheet.worksheet import Worksheet
import re
from collections import OrderedDict
from openpyxl.utils import coordinate_from_string, column_index_from_string


def RGBToHTMLColor(rgb_tuple):
    """ convert an (R, G, B) tuple to #RRGGBB 
    Code from:
       https://code.activestate.com/recipes/266466-html-colors-tofrom-rgb-tuples/history/2/ 
    """

    hexcolor = '#%02x%02x%02x' % rgb_tuple
    # that's it! '%02x' means zero-padded, 2-digit hex values
    return hexcolor
def HTMLColorToRGB(colorstring):
    """ convert #RRGGBB to an (R, G, B) tuple 
    Code from:
       https://code.activestate.com/recipes/266466-html-colors-tofrom-rgb-tuples/history/2/ 
    Value check improved with detecting hex in the color string
    """
    colorstring = colorstring.strip()
    if colorstring[0] == '#': colorstring = colorstring[1:]
    if len(colorstring) != 6:
        raise ValueError, "input #%s is not in #RRGGBB format" % colorstring
    if re.match("[a-fA-F0-9]",colorstring) == None:
        raise ValueError, "input #%s is not a hex representation of #RRGGBB " % colorstring
        
    r, g, b = colorstring[:2], colorstring[2:4], colorstring[4:]
    r, g, b = [int(n, 16) for n in (r, g, b)]
    return r, g, b

def getColors(wb):
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]

    embeddedColors = firstColorScheme.getchildren()
    colors=[]
#     for col in firstColorScheme.getchildren():
#         hexstr=col.getchildren()[0].attrib['val']
#         colors.append(hexstr)
    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        if 'window' in accent.getchildren()[0].attrib['val']:
            colors.append(accent.getchildren()[0].attrib['lastClr'])
        else:
            colors.append(accent.getchildren()[0].attrib['val'])

    return colors
def isHTMLHexColor(colorstring):
    if colorstring[0] == '#': colorstring = colorstring[1:]
    if len(colorstring) != 6:
        return False
    if re.match("[a-fA-F0-9]",colorstring) == None:
        return False
    return True

def computeLuminosity(rgbHex,tint):
    """returns RGB in 0-1 floating value tuple 
    Code obtained from :
        http://ciintelligence.blogspot.com/2012/02/converting-excel-theme-color-and-tint.html 
    Computation slightly made better.
    """
    try:
        themergb=HTMLColorToRGB(rgbHex)
    except ValueError:
        # this is not an identifiable color
        # return white
        return (1,1,1)
    hls=colorsys.rgb_to_hls(themergb[0]/255.0,themergb[1]/255.0,themergb[2]/255.0)
    if tint<0:
        l=hls[1] * (1.0 + tint)
    else:
        l= (hls[1]  + (1 - hls[1] )* tint )
    finrgb=colorsys.hls_to_rgb(hls[0],l,hls[2])
    return finrgb

def getCellColor(fg,WBCOLORS):
    if fg.type == "rgb":
        if fg.rgb[:2] == "00":
            # assuming this is transparent
            return None
        return "#"+fg.rgb[2:]
    if fg.type=="theme":
        themecolor=WBCOLORS[fg.theme]
        if isHTMLHexColor(themecolor):
            finrgb=computeLuminosity(themecolor, fg.tint)
        else:
            return None

        finhex=RGBToHTMLColor((finrgb[0]*255,finrgb[1]*255,finrgb[2]*255))
        return finhex
def wikiCellStyle(fg,bg,font=None,bold=False,italics=False,underline=False,strikethrough=False,width=None,colspan=0,rowspan=0):
    style=[]
    span=[]
    if bg==None and fg==None:
        return ""
    if bg != None:
        style.append("background-color:%s"%bg)
    if fg != None:
        style.append("color:%s"%fg)
    if font != None:
        style.append("font-family:%s"%font)
    if bold:
        style.append("font-weight:bold")
    if italics:
        style.append("font-style:italic")
    if underline:
        style.append("text-decoration:underline")
    elif strikethrough:
        style.append("text-decoration:line-through")
    if width:
        style.append("width:%s"%width)
    if colspan > 1:
        span.append("colspan=%s"%(colspan))
    if rowspan > 1:
        span.append("rowspan=%s"%(rowspan))

    return """|%s style="%s" """%(" ".join(span),  ";".join(style))
def getColRowSpan(mergeDef):
    #merge def is like A1:B2
    cells = mergeDef.split(":")
    sC,sR=coordinate_from_string(cells[0])
    sC = column_index_from_string(sC)
    eC,eR=coordinate_from_string(cells[1])
    eC = column_index_from_string(eC)
    colspan= int(eC) - int(sC) + 1
    rowspan = int(eR) - int(sR) + 1
    return colspan, rowspan

def cellToWiki(cell,WBCOLORS,ws, width=None):
    # if cell is part of merged cells and is not the first one, then skip
    colspan=0
    rowspan=0
    if cell.coordinate in ws.merged_cells:
        for mrange in ws.merged_cell_ranges:
            if mrange.startswith(cell.coordinate):
                colspan,rowspan=getColRowSpan(mrange)
                break
        else:
            return ""
    bg=getCellColor(cell.fill.fgColor, WBCOLORS)
    fg=getCellColor(cell.font.color, WBCOLORS)
    bold = cell.font.b
    italics = cell.font.i
    underline = cell.font.u
    strike=cell.font.strikethrough
    font_name= cell.font.name
    
    wikicellstr= wikiCellStyle(fg, bg,font_name,bold,italics,underline,strike,width,colspan,rowspan)
    
    if cell.value != None:
        wikicellstr+="|"+str(cell.value)+"\n"
    else:
        wikicellstr+="|\n"
    return wikicellstr    

def getColumnWidths(ws):
    colwidths=[v.width for k,v in ws.column_dimensions.iteritems()]
    return colwidths

class excelToWiki():
    
    def __init__(self,wb_name,shtnames=[],capfgcolor=None,capbgcolor=None): 
        self.wb = load_workbook(wb_name,data_only=True)
        self.sheetnames=self.wb.get_sheet_names()
        self.WBCOLORS=getColors(self.wb)
        self.wikitblmap=OrderedDict({})
        wikitbl=""
        if capfgcolor==None: capfgcolor="black"
        if capbgcolor==None: capbgcolor="#F0F0F0"

        capstyle="""style="font-weight:bold;color:%s;background-color:%s" | """%(capfgcolor,capbgcolor)
        for shtname in self.sheetnames:
            ws=self.wb.get_sheet_by_name(shtname)
            assert isinstance(ws, Worksheet)
            colwidths = getColumnWidths(ws)
            wikitbl="{|\n|+ %s %s\n"%(capstyle,shtname)
            firstrow=True
            for row in ws.iter_rows():
                if firstrow:
                    if len(row) > len(colwidths):
                        padding=len(row)-len(colwidths)
                        colwidths.extend([None for i in range(padding+1)])

                    for cell,width in zip(row,colwidths[:len(row)]):
                        wikitbl+=cellToWiki(cell, self.WBCOLORS, ws, width)
                    firstrow=False    
                else:
                    for cell in row:
                        wikitbl+= cellToWiki(cell,self.WBCOLORS, ws)
                wikitbl+="|-\n"
            wikitbl+="|}\n"
            self.wikitblmap[shtname]=wikitbl

    def getWorkbook(self):
        wikitext=""
        for k,v in self.wikitblmap.iteritems():
            wikitext+=v
        return wikitext
            
    def getSheet(self,shtname):
        if self.wikitblmap.has_key(shtname):
            return self.wikitblmap[shtname]
        else:
            return None

if __name__ == '__main__':
    e2w= excelToWiki("/Users/nnataraj/Documents/workspace/ExcelToWiki/sm-upgrade-tasks Rev 5.xlsx")
    print e2w.getSheet("Sequencing")
    
    
    