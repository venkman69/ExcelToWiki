'''
Excel to Wiki text converter

@copyright: Narayan Natarajan <venkman69@yahoo.com>
@author: venkman69
@license:
The MIT License (MIT)

Copyright (c) <year> <copyright holders>

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.
'''
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.xml.functions import fromstring,QName
from openpyxl.utils import coordinate_from_string, column_index_from_string
import colorsys
from openpyxl.worksheet.worksheet import Worksheet
import re
from collections import OrderedDict
import sys
from datetime import datetime
# use || separator instead of one cell/line

__all__=[
        "excelToWiki",
        "RGBToHTMLColor",
        "HTMLColorToRGB",
        "getWorkbookColors",
        "isHTMLHexColor",
        "getCellColor",
        "wikiStyle",
        "getColRowSpan",
        "getColumnWidths" 
         ]
def RGBToHTMLColor(rgb_tuple):
    """ convert an (R, G, B) tuple to #RRGGBB 
    Code from:
       https://code.activestate.com/recipes/266466-html-colors-tofrom-rgb-tuples/history/2/ 
    """

    hexcolor = '#%02x%02x%02x' % rgb_tuple
    # that's it! '%02x' means zero-padded, 2-digit hex values
    return hexcolor
def HTMLColorToRGB(colorstring):
    """ convert #RRGGBB to an (R, G, B) tuple 
    Code from:
       https://code.activestate.com/recipes/266466-html-colors-tofrom-rgb-tuples/history/2/ 
    Value check improved with detecting hex in the color string
    """
    colorstring = colorstring.strip()
    if colorstring[0] == '#': colorstring = colorstring[1:]
    if len(colorstring) != 6:
        raise ValueError, "input #%s is not in #RRGGBB format" % colorstring
    if re.match("[a-fA-F0-9]",colorstring) == None:
        raise ValueError, "input #%s is not a hex representation of #RRGGBB " % colorstring
        
    r, g, b = colorstring[:2], colorstring[2:4], colorstring[4:]
    r, g, b = [int(n, 16) for n in (r, g, b)]
    return r, g, b

def getWorkbookColors(wb):
    """gets themed colors from the workbook by inspecting the xml 
      Code obtained from: 
          https://groups.google.com/forum/#!topic/openpyxl-users/v2FDsbDDTqU
      """
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]

#     embeddedColors = firstColorScheme.getchildren()
    colors=[]
#     for col in firstColorScheme.getchildren():
#         hexstr=col.getchildren()[0].attrib['val']
#         colors.append(hexstr)
    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6','hlink', 'folHlink']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        if accent == None:
            continue
        if 'window' in accent.getchildren()[0].attrib['val']:
            colors.append(accent.getchildren()[0].attrib['lastClr'])
        else:
            colors.append(accent.getchildren()[0].attrib['val'])

    return colors
def isHTMLHexColor(colorstring):
    if colorstring[0] == '#': colorstring = colorstring[1:]
    if len(colorstring) != 6:
        return False
    if re.match("[a-fA-F0-9]",colorstring) == None:
        return False
    return True

def computeLuminosity(rgbHex,tint):
    """returns RGB in 0-1 floating value tuple 
    Code obtained from :
        http://ciintelligence.blogspot.com/2012/02/converting-excel-theme-color-and-tint.html 
    Computation slightly made better.
    """
    try:
        themergb=HTMLColorToRGB(rgbHex)
    except ValueError:
        # this is not an identifiable color
        # return white
        return (1,1,1)
    hls=colorsys.rgb_to_hls(themergb[0]/255.0,themergb[1]/255.0,themergb[2]/255.0)
    if tint<0:
        l=hls[1] * (1.0 + tint)
    else:
        l= (hls[1]  + (1 - hls[1] )* tint )
    finrgb=colorsys.hls_to_rgb(hls[0],l,hls[2])
    return finrgb

def getCellColor(fg,WBCOLORS):
    if fg.type == "rgb":
        if fg.rgb[:2] == "00":
            # assuming this is transparent
            return None
        return "#"+fg.rgb[2:]
    if fg.type=="theme":
        themecolor=WBCOLORS[fg.theme]
        if isHTMLHexColor(themecolor):
            finrgb=computeLuminosity(themecolor, fg.tint)
        else:
            return None

        finhex=RGBToHTMLColor((finrgb[0]*255,finrgb[1]*255,finrgb[2]*255))
        return finhex
def wikiStyle(style,passthrough=None):
# def wikiStyle(stylefg,bg,font=None,bold=False,italics=False,underline=False,strikethrough=False,width=None,colspan=0,rowspan=0):
    """ passthrough is a list of uninspected style items"""
    span=[]
    resstyle=[]
    if style.has_key("bg") and style["bg"] != None:
        resstyle.append("background-color:%s"%style["bg"])
    if style.has_key("fg") and style["fg"] != None:
        resstyle.append("color:%s"%style["fg"])
    if style.has_key("font_name"):
        resstyle.append("font-family:%s"%style["font_name"])
    if style.has_key("bold") and style["bold"]:
        resstyle.append("font-weight:bold")
    if style.has_key("italics") and style["italics"]:
        resstyle.append("font-style:italic")
    if style.has_key("underline") and style["underline"]:
        resstyle.append("text-decoration:underline")
    elif style.has_key("strike") and style["strike"]:
        resstyle.append("text-decoration:line-through")
    if style.has_key("width") and style['width'] != None:
        resstyle.append("width:%sin"%(style["width"]/12))
    if style.has_key("colspan") and style["colspan"] !=None:
        span.append("colspan=%s"%(style["colspan"]))
    if style.has_key("rowspan") and style["rowspan"] !=None:
        span.append("rowspan=%s"%(style["rowspan"]))
    if style.has_key("align") and style["align"] != None:
        align="""align="%s" """%style["align"]
    else:
        align=""
    if passthrough:
        resstyle.extend(passthrough)
    if len(span) == 0 and len(resstyle)==0:
        return ""
    return """%s style="%s" %s """%(" ".join(span),  ";".join(resstyle),align)

def getColRowSpan(mergeDef):
    #merge def is like A1:B2
    cells = mergeDef.split(":")
    sC,sR=coordinate_from_string(cells[0])
    sC = column_index_from_string(sC)
    eC,eR=coordinate_from_string(cells[1])
    eC = column_index_from_string(eC)
    colspan= int(eC) - int(sC) + 1
    rowspan = int(eR) - int(sR) + 1
    if colspan == 1:
        colspan=None
    if rowspan == 1:
        rowspan=None
    return colspan, rowspan


class wikiCell():
    wikicellstr=None
    bg=None
    fg=None
    underline=None
    strike=None
    font_name=None
    bold=None
    value=None
    style=None
    merged=False
    def __init__(self,cell,WBCOLORS,ws):
        colspan=None
        rowspan=None
        self.col,self.rownum=coordinate_from_string(cell.coordinate)
        self.style={}
        self.width=None
#         if rownum == 1:
#             if colwidths.has_key(col):
#                 self.width=colwidths[col]
            
        if cell.coordinate in ws.merged_cells:
            for mrange in ws.merged_cell_ranges:
                if mrange.startswith(cell.coordinate):
                    colspan,rowspan=getColRowSpan(mrange)
                    break
            else:
                self.merged=True
                return
                
        self.cell = cell
        cval=""
            
        try:
            if cell.value == None:
                cval="&nbsp;"
            else:
                if isinstance(cell.value, unicode):
                    cval=cell.value.strip()
                if cell.is_date:
                    cval=self.__doDateFmt()
                elif cell.number_format.endswith("%"):
                    dotind = cell.number_format.find(".")
                    if dotind == -1:
                        fmtStr = "%0d%%"
                    else:
                        fmtStr = "%" + str(dotind) + ".%d"%(len(cell.number_format) - dotind - 2) + "f%%"
                    cval = unicode(fmtStr % (cell.value * 100),'utf-8')
                else:
                    cval=unicode(str(cell.value),'utf-8').strip()
        except:
                cval="" # not sure what to do here
            
        # if value starts with http, then wrap it in []s,otherwise it is used verbatim
        if cval.startswith("http://"):
            cval="["+cval+"]"
        
       # cval=cval.replace(u"\n", "<br/>")
        
        self.value = cval
        self.bg=getCellColor(cell.fill.fgColor, WBCOLORS)
        self.fg=getCellColor(cell.font.color, WBCOLORS)
        if self.fg== "#000000":
            self.fg=None
        self.bold = cell.font.b
        self.italics = cell.font.i
        self.underline = cell.font.u
        self.strike=cell.font.strikethrough
        self.font_name= cell.font.name

        self.style["bg"]=self.bg
        self.style["fg"]=self.fg
        self.style["underline"]=self.underline
        self.style["strike"]=self.strike
        self.style["italics"]=self.italics
        self.style["font_name"]=self.font_name
        self.style["bold"]=self.bold
        self.style["colspan"]=colspan
        self.style["rowspan"]=rowspan
        self.style["width"]=self.width
        self.style["align"]=self.cell.alignment.horizontal
    def __doDateFmt(self):
        """Returns string representation for a few fixed formats of excel dates"""
        # support a few popular formats but not everything
        # since parsing m for month vs m for min seems unclear.
        #[$-409]m/d/yy\ h:mm\ AM/PM;@
        #m/d/yy\ h:mm;@
        #[$-409]mmm\-yy;@
        #[$-409]mmmm\-yy;@
        #m/d;@
        # [$-409] is a locale string - we will assume english for now
        # the @ implies 'as is'
        strfstr=""
        strftimemap={
                 "h":"%H", #one or more 'h' is still hours
                 "h:mm":"%H:%M",
                 "hh:mm":"%H:%M",
                 "h:m":"%H:%M",
                 "AM/PM":"%p",
                 }
        strfdatemap={
                 "ddd":"%a",
                 "d":"%d",
                 "dd":"%d",
                 "mmmm":"%B",
                 "mmm":"%b",
                 "mm":"%m",
                 "m":"%m",
                 "yyyy":"%Y",
                 "yy":"%y",
                 "y":"%y",
                 "s":"%S",
                 "ss":"%S",
                 "\\ ":" ",
                 "[$-":"",
                 "409":"",
                 "]":"",
                 ";@":"",
                 "\\,\\":",",
                 "\\-":"-",
                 "\\- ":"- ",
                 "\\,\\ ":", ",
                 }
        # find AM/PM
        AMPM=False
        if "AM/PM" in self.cell.number_format:
            AMPM=True
        fmt=False
        for tok in re.split(r'([dmyh:]+|[ ]+|\W+)',self.cell.number_format):
            if strfdatemap.has_key(tok):
                fmt=True
                strfstr+=strfdatemap[tok]
            elif strftimemap.has_key(tok):
                fmt=True
                if AMPM:
                    strfstr+=strftimemap[tok].replace("H","I")+"%p"
                else:
                    strfstr+=strftimemap[tok]
            else:
                strfstr+=tok
        if fmt:
            return self.cell.value.strftime(strfstr)
        # if all else fails return a string rep of the entire datetime default
        return str(self.cell.value)

    def getWikiStr(self,rowstyle=[],colwidths=None): 
        if self.merged:
            return ""
        cellstyle={}
        for style in self.style:
            if style not in rowstyle:
                cellstyle[style]=self.style[style]
        if colwidths!=None and colwidths.has_key(self.col) and colwidths[self.col] != None:
            cellstyle["width"]=colwidths[self.col]
            
        wikiCellStyle= wikiStyle(cellstyle)
        
        #if self.value != None and self.value != "":
        if not "\n" in self.value:
            if self.merged:
                return ""
            wikicellstr=""
            if wikiCellStyle != "":
                wikicellstr += wikiCellStyle + "| "
            wikicellstr+= self.value

        else:
            wikicellstr=""
            if wikiCellStyle != "":
                wikicellstr += wikiCellStyle + "|"
            wikicellstr+="\n" + self.value+"\n"
#         else:
#             wikicellstr="|\n"
        return wikicellstr    

class wikiRow():
    style=None
    rowwiki=""
    def __init__(self,row, WBCOLORS, ws,preserve_width=True):
        celllist=[]
        styleList=[]
        for cell in row:
            wcell=wikiCell(cell,WBCOLORS,ws)
            celllist.append(wcell)
            styleList.append(wcell.style)
        # resolve common styles
        self.style = commonStyle(styleList)
        col,rownum=coordinate_from_string(cell.coordinate)
        if preserve_width ==True and rownum==1:
            colwidths = getColumnWidths(ws)
            if colwidths.has_key(col):
                width=colwidths[col]
        else:
            colwidths=None

        self.rowwiki="| "
        cellList=[]
        for cell in celllist:
            if not cell.merged:
                cellList.append(cell.getWikiStr(self.style.keys(),colwidths))
        # also handle the special case of using single pipe when a return character is encountered
        self.rowwiki+=("|| ".join(cellList)+"\n").replace("\n||","\n|") 

    def getWikiStr(self,tblstyle=[]): 
        rowstyle={}
            
        for styleItem in self.style:
            if styleItem not in tblstyle:
                rowstyle[styleItem]=self.style[styleItem]
        wikirowstr= "|-%s\n"%wikiStyle(rowstyle)
        wikirowstr+=self.rowwiki
        return wikirowstr
    def __repr__(self):
        return self.getWikiStr()

class wikiTbl():
    style=None
    tblwiki=""
    rowList=None
    def __init__(self,ws,WBCOLORS,capstyle,preserve_width=True):
        assert isinstance(ws, Worksheet)
        self.shtname=ws.title
        self.capstyle=capstyle
        wikitbl="""{|border=1 style="border-collapse: collapse;border-color:#aaaaaa"\n|+ %s %s\n"""%(capstyle,self.shtname)
        self.rowList=[]
        styleList=[]
        for row in ws.iter_rows():
            wrow=wikiRow(row, WBCOLORS, ws,preserve_width)
            self.rowList.append(wrow)
            styleList.append(wrow.style)
        self.style = commonStyle(styleList)

    def getWikiStr(self):
        if len(self.rowList)==0:
            return ""
        wikitblstr="{| border=1 %s\n" % wikiStyle(self.style,["border-collapse: collapse; border-color: #aaaaaa"]) 
        wikitblstr+="|+%s %s\n"%(self.capstyle,self.shtname)
        for row in self.rowList:
            wikitblstr+=row.getWikiStr(self.style.keys())
        wikitblstr+="|}"
        return wikitblstr

def commonStyle(styleList):
    comstyle={}
    for style in styleList:
        for key,val in style.iteritems():
            if comstyle.has_key(key):
                comstyle[key].append(val)
            else:
                comstyle[key]=[val]
    cunique={}
    for ck,cv in comstyle.iteritems():
        scv = list(set(cv))
        if len(cv) == len(styleList) and len(scv)==1 and scv[0] != None and scv[0] != False and scv[0] != 0:
            cunique[ck]=scv[0]
    return cunique


def getColumnWidths(ws):
    """returns column widths of a worksheet"""
    colwidths={k:v.width for k,v in ws.column_dimensions.iteritems()}
    return colwidths

class excelToWiki():
    wb=None
    sheetnames=None
    wikitblmap=None
    WBCOLORS=None
     
    def __init__(self,wb,shtnames=[],capfgcolor=None,capbgcolor=None,preserve_width=True): 
        """
           wb can be a name/path to excel file or a file like object
           shtnames: <list> Specify sheet names to convert
           capfgcolor: caption foreground color in html hex color format (eg: '#FF0000' or 'red') 
           capbgcolor: caption background color in html hex color format (eg: '#00FF00' or 'yellow') 
        """
        try:
            self.wb = load_workbook(wb,data_only=True)
        except:
            print sys.exc_info()
            raise Exception("Could not load excel workbook")
        if len(shtnames)>0:
            self.sheetnames=shtnames
        else:
            self.sheetnames=self.wb.get_sheet_names()
        self.WBCOLORS=getWorkbookColors(self.wb)
        self.wikitblmap=OrderedDict({})
        wikitbl=""
        if capfgcolor==None: capfgcolor="black"
        if capbgcolor==None: capbgcolor="#F0F0F0"

        capstyle="""style="font-weight:bold;color:%s;background-color:%s" | """%(capfgcolor,capbgcolor)
        for shtname in self.sheetnames:
            ws=self.wb.get_sheet_by_name(shtname)
            assert isinstance(ws, Worksheet)
            if ws == None:
                continue
            wt=wikiTbl(ws, self.WBCOLORS, capstyle,preserve_width)
            self.wikitblmap[shtname]=wt.getWikiStr()

    def getWorkbook(self):
        wikitext=""
        for k,v in self.wikitblmap.iteritems():
            wikitext+=v+"\n"
        return wikitext
            
    def getSheet(self,shtname):
        if self.wikitblmap.has_key(shtname):
            return self.wikitblmap[shtname]
        else:
            return None
